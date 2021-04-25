#Update daily cryptocurrencies
from requests import Request, Session, request
import requests
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
import json
import pickle
import pandas as pd
import numpy as np
import openpyxl
import os
from bs4 import BeautifulSoup
import datetime

#get current USD-Price
url = "https://www.x-rates.com/calculator/?from=EUR&to=USD&amount=1"
html_content= requests.get(url).text

soup = BeautifulSoup(html_content, "lxml")
usd_container = soup.findAll('span', {"class":"ccOutputRslt"})
usd_ratio = usd_container[0].text

#API Capcoinmarket
url = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest'
parameters = {
  'start':'1',
  'limit':'100',
  'convert':'USD'
}
headers = {
  'Accepts': 'application/json',
  'X-CMC_PRO_API_KEY': '3559a0f0-81a8-4d39-b209-e66bdc3119e7',
}

session = Session()
session.headers.update(headers)

try:
  response = session.get(url, params = parameters)
  data = json.loads(response.text)
  print(data)
  #data = response.json()
  
except (ConnectionError, Timeout, TooManyRedirects) as e:
  print(e)

#For testing i saved the Dictionary into an pickle object to save my API-Key-Credits
with open('crypto_data1.pkl', 'wb') as crypto_dict:
    pickle.dump(data, crypto_dict)
# with open('crypto_data1.pkl', 'rb') as crypto_dict:
#     data = pickle.load(crypto_dict)
# crypto_dict.close()

coin_list_name = []
coin_list_price = []

for i in range(0,99):
    crypto_name = data["data"][i]["name"]
    crypto_price = data["data"][i]["quote"]["USD"]["price"]
    #crypto_time= data["status"]["timestamp"]

    coin_list_name.append(crypto_name)
    coin_list_price.append(crypto_price)

    
#building the appropriate Dictioanry to read into Dataframe
coin_dict = {
    "Cryptocurrency": coin_list_name,
    "Price in USD": coin_list_price,
    "Date": datetime.datetime.today().date()
            }

#building Dataframe            
coin_df = pd.DataFrame.from_dict(coin_dict)
coin_df.set_index(["Cryptocurrency", "Price in USD"])
coin_df["Date"]=coin_df["Date"].astype(str)

date_changer = lambda row: datetime.datetime.strptime(row["Date"], "%Y-%m-%d")

try:
  # changing format of Date Column
  coin_df["Date"] = coin_df.apply(date_changer,axis=1)
  coin_df["Date"] = coin_df["Date"].dt.strftime("%d.%m.%Y")
  
  #saving File with appropriate Filename
  date = datetime.date.today().strftime("%Y_%m_%d")
  csv_file_name = date + "_Dashboard_Crypto.csv"
  coin_df.to_csv(r"C:\Users\bennk\Documents\Programmierung\Pythonprogramms\Crypto_pro\Price_Track\\"+csv_file_name,index=None, sep=";")
  
  filename_read = r"C:\Users\bennk\Documents\Programmierung\Pythonprogramms\Crypto_pro\Dashboard_Crypto.xlsx"
  #writing file
  writer = pd.ExcelWriter(filename_read, engine="openpyxl", mode="w",index=False)
  writer.book = openpyxl.load_workbook(filename_read)
  wb = openpyxl.Workbook(filename_read)   #CApital Letter!!!
  writer.sheets = dict((ws.title,ws) for ws in writer.book.worksheets)
  coin_df.to_excel(writer, sheet_name="Crypto_Price",startcol = 0, index=False)
  writer.save()
  writer.close()

except FileNotFoundError:
  print("Use the preperated file in the right folder")

wb = openpyxl.load_workbook(filename_read)
sheet = wb["Crypto_Price"]
sheet1 = wb["Crypto_Omni"]
sheet1["I2"] = "usd_ratio"
sheet.column_dimensions["A"].width = 20
sheet.column_dimensions["B"].width = 20
sheet.column_dimensions["C"].width = 20

wb.save(filename_read)



